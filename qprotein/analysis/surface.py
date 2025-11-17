"""
# ------------------------------------------------------------------------------
# Author:    Dou Zhixin
# Email:     bj600800@gmail.com
# DATE:      2024/11/27

# Description: Run code for surface calculation
# ------------------------------------------------------------------------------
"""
import os
import configparser

from openpyxl import Workbook
from openpyxl.styles import Alignment

from qprotein.feature import compute_surface
from qprotein.utilities import logger

logger = logger.setup_log(name=__name__)


def save2xlsx(data, xlsx_file):
    workbook = Workbook()

    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    statistics_sheet = workbook.create_sheet("Statistics")
    detail_sheet = workbook.create_sheet("Details")

    statistics_sheet['A1'] = 'Enzyme names'
    statistics_sheet['B1'] = 'Negative percentage'
    statistics_sheet['C1'] = 'Positive percentage'
    statistics_sheet['D1'] = 'Net charges'

    for row_index, row_perc in enumerate(data, start=2):
        statistics_sheet.cell(row=row_index, column=1, value=row_perc[0])
        statistics_sheet.cell(row=row_index, column=2, value=row_perc[2]['negative'])
        statistics_sheet.cell(row=row_index, column=3, value=row_perc[2]['positive'])
        statistics_sheet.cell(row=row_index, column=4, value=row_perc[3])

    # Write data into the sheet
    for row_index, residue in enumerate(data, start=1):
        name = residue[0]

        start_row = row_index * 2 - 1
        end_row = start_row + 1

        detail_sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        detail_sheet.cell(row=start_row, column=1, value=name)
        detail_sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        detail_sheet.cell(row=row_index * 2 - 1, column=1, value=name + '_chainA')

        # Write residue indices and types
        for col_index, (res_idx, res_type) in enumerate(residue[1], start=2):
            detail_sheet.cell(row=row_index * 2 - 1, column=col_index, value=res_idx)
            detail_sheet.cell(row=row_index * 2, column=col_index, value=res_type)

    for row in statistics_sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in detail_sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    workbook.save(xlsx_file)
    workbook.close()
    logger.info('Results saved to {}'.format(xlsx_file))


def run(work_dir, structure_dir, config_file):
    #### CONFIGURATION PARSER ####
    config = configparser.ConfigParser()
    config.read(config_file)
    dssp_bin = config.get('binary', 'dssp')
    apbs_bin = config.get('binary', 'apbs')
    pdb2pqr_bin = config.get('binary', 'pdb2pqr')
    #### END OF CONFIGURATION PARSER ####

    xlsx_file = os.path.join(work_dir, "surface.xlsx")
    data = compute_surface.run(structure_dir, dssp_bin=dssp_bin, pdb2pqr_bin=pdb2pqr_bin, apbs_bin=apbs_bin)
    save2xlsx(data, xlsx_file)

if __name__ == '__main__':
    structure_dir = r"/Users/douzhixin/Developer/qProtein/qProtein-main/test/structure"
    xlsx_file = r"/Users/douzhixin/Developer/qProtein/qProtein-main/test/surface.xlsx"
    config_file = r"/Users/douzhixin/Developer/qProtein/qProtein-main/config.ini"
    run(structure_dir, xlsx_file, config_file)