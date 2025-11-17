"""
# ------------------------------------------------------------------------------
# Author:    Dou Zhixin
# Email:     bj600800@gmail.com
# DATE:      2024/11/27

# Description: run code for interaction
# ------------------------------------------------------------------------------
"""
import os
from pathlib import Path

import biotite.structure.io as strucio
from openpyxl import Workbook
from openpyxl.styles import Alignment
from tqdm import tqdm

from qprotein.feature import hydrophobic, hbond, salt_bridge, disulfide_bond
from qprotein.utilities import logger

logger = logger.setup_log(name=__name__)

def calc_interactions(structure_dir):
    interactions = {}
    pdb_files = list(Path(structure_dir).glob("*.pdb"))
    for structure_path in tqdm(pdb_files, desc='Computing interactions for monomers'):
        _interaction = {}
        file_name = structure_path.stem
        structure = strucio.load_structure(structure_path)
        hydrophobic_ret = hydrophobic.run(structure)
        _interaction['hydrophobic'] = hydrophobic_ret
        hbond_ret = hbond.run(structure_path, 'pairs')
        _interaction['hbond'] = hbond_ret
        salt_bridge_ret = salt_bridge.run(structure, 'pairs')
        _interaction['salt_bridge'] = salt_bridge_ret
        disulfide_bond_ret = disulfide_bond.run(structure, 'pairs')
        _interaction['disulfide_bond'] = disulfide_bond_ret
        interactions[file_name] = _interaction
    return interactions


def save2xlsx(data_dict, xlsx_file):
    workbook = Workbook()
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    sheets = []
    sheet_names = ["hydrophobic", "hbond", "salt_bridge", "disulfide_bond"]
    for name in sheet_names:
        sheets.append(workbook.create_sheet(name))

    for protein_name, interaction in data_dict.items():
        for interact_name, value in interaction.items():
            if interact_name == "hydrophobic":
                sheet = workbook["hydrophobic"]
                max_row = sheet.max_row
                if max_row == 1:
                    sheet.cell(row=1, column=1, value='protein names')
                    sheet.column_dimensions['A'].width = len('protein names') * 2
                    sheet.cell(row=1, column=2, value='cluster name')
                    sheet.column_dimensions['B'].width = len('cluster name') * 2
                    sheet.cell(row=1, column=3, value='area(A^2)')
                    sheet.column_dimensions['C'].width = len('area(A^2)') * 2
                    sheet.cell(row=1, column=4, value='residues')
                    sheet.column_dimensions['D'].width = len('residues') * 2
                for i, (cluster_idx, info) in enumerate(value.items()):
                    start_row = max_row + 1
                    end_row = start_row + len(value) - 1
                    merge_range = f'A{start_row}:A{end_row}'  # column A
                    sheet.merge_cells(merge_range)
                    cell = sheet[merge_range.split(':')[0]]
                    cell.value = protein_name
                    sheet.cell(row=start_row + i, column=2, value=cluster_idx)
                    sheet.cell(row=start_row + i, column=3, value=info[0])
                    sheet.cell(row=start_row + i, column=4, value='+'.join([str(_) for _ in info[1]]))
            else:
                sheet = workbook[interact_name]
                max_column = sheet.max_column
                if max_column == 1:
                    start_column = max_column
                else:
                    start_column = max_column + 1
                end_column = start_column + 1
                sheet.cell(row=1, column=start_column, value=protein_name)
                sheet.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=end_column)
                for i, pair in enumerate(value):
                    sheet.cell(row=i + 2, column=start_column, value=pair[0][1])
                    sheet.cell(row=i + 2, column=end_column, value=pair[1][1])

    for sh in sheets:
        for row in sh.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    workbook.save(xlsx_file)
    workbook.close()
    logger.info('Results saved to directory: {}'.format(xlsx_file))


def generate_pml_file(data_dict, pdb_dir, pml_dir):
    for name, four_interaction in data_dict.items():
        output_pml = os.path.join(pml_dir, f"{name}.pml")
        pdb_path = os.path.join(pdb_dir, name + '.pdb')
        with open(output_pml, 'w') as pml_file:
            pml_file.write(f"load {pdb_path}, {name}\n")
            pml_file.write("remove solvent\n")
            pml_file.write(f"color white, {name}\n")

            # pymol scripts for each hydrophobic cluster
            if four_interaction['hydrophobic']:
                for i, (cluster_name, cluster_info) in enumerate(four_interaction['hydrophobic'].items()):
                    select_str = f"select {cluster_name}, res { '+'.join(map(str, cluster_info[1])) }"
                    pml_file.write(select_str + '\n')
                    pml_file.write(f"show spheres, {cluster_name}\n")
                    pml_file.write(f"color {color_iter(i)}, {cluster_name}\n")

            point_interaction = ['salt_bridge', 'disulfide_bond']
            for point_i in point_interaction:
                if four_interaction[point_i]:
                    res_list = []
                    for pair in four_interaction[point_i]:
                        distance_str = f'distance dist_{point_i}'
                        for res in pair:
                            distance_str += f", id {str(res[0])}"
                            res_list.append(str(res[1]))
                        pml_file.write(distance_str + '\n')
                    sele_str = f"select {point_i}, res {'+'.join(res_list)}"
                    pml_file.write(sele_str + '\n')
                    pml_file.write(f'show sticks, {point_i}\n')
                    pml_file.write(f'color atomic, {point_i}\n')
            pml_file.write(f'remove hydrogen\n')
    logger.info('Visualization pml files saved to {}'.format(pml_dir))


def color_iter(idx):
    colors = ['wheat', 'palegreen', 'lightblue', 'paleyellow', 'lightpink', 'palecyan', 'lightorange', 'bluewhite',
              'tv_red', 'tv_green', 'tv_blue', 'tv_yellow', 'lightmagenta', 'tv_orange']
    return colors[idx % len(colors)]


def run(work_dir, pdb_dir, pml=False):
    data_dict = calc_interactions(pdb_dir)
    xlsx_file = os.path.join(work_dir, "visual.xlsx")
    save2xlsx(data_dict, xlsx_file)
    if pml:
        pml_dir = os.path.join(work_dir, "pml")
        if not os.path.exists(pml_dir):
            os.mkdir(pml_dir)
        generate_pml_file(data_dict, pdb_dir, pml_dir)
